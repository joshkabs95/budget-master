"""
Parser Excel BudgetMaster v2.
Lit jusqu'à 6 feuilles et retourne un dict enrichi :
  {
    transactions : [...],
    categories   : [...],   # [{name, bucket, icon}]
    goals        : [...],   # [{name, icon, current_amount, target_amount, deadline}]
    scores       : {...},   # {category_name: score_int}
  }
"""
import calendar
import re
from datetime import date


_OCC_DAYS = {1: [15], 2: [8, 22], 3: [7, 14, 21], 4: [7, 14, 21, 28]}

_SKIP_LABELS = frozenset([
    "", "none", "nan", "libellé", "libelle", "source", "label",
    "total", "totaux", "objectif", "catégorie", "categorie", "nom",
])


# ── Public entry point ────────────────────────────────────────────────────────

def parse_excel(path: str) -> dict:
    """
    Parse an xlsx file. Returns dict with transactions/categories/goals/scores.
    Falls back to empty lists on any error.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return _empty()

    try:
        wb = load_workbook(path, data_only=True)
    except Exception:
        return _empty()

    year, month = _read_config(wb)
    if not year:
        year = date.today().year
    if not month:
        month = date.today().month

    max_day = calendar.monthrange(year, month)[1]

    return {
        "transactions": _parse_transactions(wb, year, month, max_day),
        "categories":   _parse_categories(wb),
        "goals":        _parse_goals(wb),
        "scores":       _parse_scores(wb),
    }


def _empty():
    return {"transactions": [], "categories": [], "goals": [], "scores": {}}


# ── Transactions ──────────────────────────────────────────────────────────────

def _parse_transactions(wb, year, month, max_day):
    transactions = []

    ws_rev = _find_sheet(wb, ["revenus", "revenu", "revenue"])
    if ws_rev:
        for row in ws_rev.iter_rows(min_row=3, values_only=True):
            label  = _clean_str(row[0] if len(row) > 0 else None)
            amount = _clean_amount(row[1] if len(row) > 1 else None)
            day    = _clean_day(row[2] if len(row) > 2 else None, max_day)
            if not label or not amount or amount <= 0:
                continue
            transactions.append(_make(label, amount, date(year, month, day), is_credit=True))

    ws_fix = _find_sheet(wb, ["fixes", "fixe", "fixed", "dépenses fixes", "depenses fixes"])
    if ws_fix:
        for row in ws_fix.iter_rows(min_row=3, values_only=True):
            if len(row) < 2:
                continue
            label    = _clean_str(row[0])
            amount   = _clean_amount(row[1])
            category = _clean_str(row[2]) if len(row) > 2 else None
            day      = _clean_day(row[3] if len(row) > 3 else None, max_day)
            if not label or not amount or amount <= 0:
                continue
            transactions.append(_make(label, amount, date(year, month, day),
                                      is_credit=False, category=category))

    ws_var = _find_sheet(wb, ["variables", "variable", "dép. variables", "dep. variables"])
    if ws_var:
        for row in ws_var.iter_rows(min_row=3, values_only=True):
            if len(row) < 3:
                continue
            label    = _clean_str(row[0])
            min_amt  = _clean_amount(row[1])
            max_amt  = _clean_amount(row[2])
            category = _clean_str(row[3]) if len(row) > 3 else None
            occ      = _clean_int(row[4]) if len(row) > 4 else 1
            if not label or min_amt is None or max_amt is None:
                continue
            if min_amt <= 0 and max_amt <= 0:
                continue
            avg = round((min_amt + max_amt) / 2, 2)
            occ = max(1, min(occ or 1, 4))
            for d in _OCC_DAYS.get(occ, [15]):
                transactions.append(_make(label, avg, date(year, month, min(d, max_day)),
                                          is_credit=False, category=category))

    return transactions


# ── Categories ────────────────────────────────────────────────────────────────

def _parse_categories(wb) -> list:
    ws = _find_sheet(wb, ["catégories", "categories", "categorie"])
    if not ws:
        return []

    results = []
    valid_buckets = {"needs", "wants", "savings"}

    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row:
            continue
        name   = _clean_str(row[0])
        bucket = str(row[1] or "").strip().lower() if len(row) > 1 else ""
        icon   = str(row[2] or "").strip()         if len(row) > 2 else ""
        if not name or bucket not in valid_buckets:
            continue
        results.append({
            "name":   name,
            "bucket": bucket,
            "icon":   icon or "📁",
        })

    return results


# ── Goals ─────────────────────────────────────────────────────────────────────

def _parse_goals(wb) -> list:
    ws = _find_sheet(wb, ["objectifs", "objectif", "goals", "goal"])
    if not ws:
        return []

    results = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row:
            continue
        name    = _clean_str(row[0])
        icon    = str(row[1] or "🎯").strip()         if len(row) > 1 else "🎯"
        current = _clean_amount(row[2])                if len(row) > 2 else 0.0
        target  = _clean_amount(row[3])                if len(row) > 3 else None
        deadline = _clean_date(row[5])                 if len(row) > 5 else None
        if not name or not target or target <= 0:
            continue
        results.append({
            "name":           name,
            "icon":           icon or "🎯",
            "current_amount": current or 0.0,
            "target_amount":  target,
            "deadline":       deadline or "",
            "type":           "savings",
        })

    return results


# ── Scores wants ──────────────────────────────────────────────────────────────

def _parse_scores(wb) -> dict:
    ws = _find_sheet(wb, ["scores", "wants", "pertinence"])
    if not ws:
        return {}

    scores = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row:
            continue
        name  = _clean_str(row[0])
        score = _clean_int(row[1]) if len(row) > 1 else None
        if not name or score is None:
            continue
        score = max(1, min(5, score))
        scores[name] = score

    return scores


# ── Config ────────────────────────────────────────────────────────────────────

def _read_config(wb) -> tuple:
    if "_config" in wb.sheetnames:
        ws = wb["_config"]
        try:
            raw = str(ws["A2"].value or "")
            if "-" in raw:
                y, m = raw.split("-")
                return int(y), int(m)
        except Exception:
            pass
    # Fallback: scan sheet titles for a year
    for name in wb.sheetnames:
        ws = wb[name]
        cell = str(ws["A1"].value or "")
        m = re.search(r"(\d{4})", cell)
        if m:
            return int(m.group(1)), date.today().month
    return None, None


# ── Sheet finder ──────────────────────────────────────────────────────────────

def _find_sheet(wb, keywords):
    for name in wb.sheetnames:
        nl = name.lower().strip()
        if any(kw in nl for kw in keywords):
            return wb[name]
    return None


# ── Helpers ───────────────────────────────────────────────────────────────────

def _clean_str(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    if s.lower() in _SKIP_LABELS:
        return None
    return s or None


def _clean_amount(val) -> float | None:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace("€", "").replace("\xa0", "").replace(" ", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _clean_day(val, max_day: int) -> int:
    if val is None:
        return min(15, max_day)
    try:
        d = int(float(str(val)))
        return max(1, min(d, max_day))
    except (ValueError, TypeError):
        return min(15, max_day)


def _clean_int(val) -> int | None:
    if val is None:
        return None
    try:
        return int(float(str(val)))
    except (ValueError, TypeError):
        return None


def _clean_date(val) -> str | None:
    if val is None:
        return None
    if hasattr(val, "strftime"):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            from datetime import datetime
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def _make(label: str, amount: float, txn_date, is_credit: bool, category: str | None = None) -> dict:
    from apps.transactions.services.categorization import categorize_label, compute_import_hash
    if category:
        cat_name   = category
        confidence = 100
    else:
        cat = categorize_label(label)
        cat_name   = cat["name"] if cat else "Autre"
        confidence = cat["confidence"] if cat else 0

    import_hash = compute_import_hash(str(txn_date), str(round(amount, 2)), label)
    return {
        "date":               str(txn_date),
        "amount":             amount,
        "label":              label,
        "type":               "income" if is_credit else "expense",
        "suggested_category": cat_name,
        "confidence":         confidence,
        "import_hash":        import_hash,
        "is_duplicate":       False,
    }
