"""
Génération du template Excel BudgetMaster v2 — 6 feuilles.
"""
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Palette ───────────────────────────────────────────────────────────────────
BG_PAGE     = "080F1C"
BG_HEADER   = "0B1A2E"
BG_ROW      = "0D1B2E"
BG_ROW_ALT  = "0F2035"
BG_EXAMPLE  = "17263D"
BG_TOTAL    = "050C17"

FG_ACCENT   = "00D4AA"   # teal
FG_GREEN    = "34D399"
FG_RED      = "F87171"
FG_ORANGE   = "FBBF24"
FG_BLUE     = "60A5FA"
FG_PURPLE   = "A78BFA"
FG_GOLD     = "FCD34D"
FG_TEXT     = "E2E8F0"
FG_MUTED    = "64748B"
FG_DIM      = "334155"
FG_EXAMPLE  = "3D5270"

_MONTHS_FR = [
    "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
    "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre",
]

BUCKET_CHOICES = ["needs", "wants", "savings"]

DEFAULT_CATEGORIES = [
    ("Logement",      "needs",    "🏠"),
    ("Énergie",       "needs",    "⚡"),
    ("Crédit",        "needs",    "💳"),
    ("Alimentation",  "needs",    "🛒"),
    ("Transport",     "needs",    "⛽"),
    ("Santé",         "needs",    "💊"),
    ("Loisirs",       "wants",    "🎬"),
    ("Sorties",       "wants",    "🍽️"),
    ("Shopping",      "wants",    "👗"),
    ("Abonnements",   "wants",    "📱"),
    ("Bien-être",     "wants",    "🧘"),
    ("Voyages",       "wants",    "✈️"),
    ("Épargne",       "savings",  "💰"),
    ("Investissement","savings",  "📈"),
]

DEFAULT_GOALS = [
    ("Voyage Miami",    "✈️",  0,  2000,  "", ""),
    ("Fond d'urgence",  "🛡️",  0,  5000,  "", "Livret A"),
    ("Voyage Congo",    "🌍",  0,  2500,  "", ""),
    ("Business perso",  "🚀",  0,  10000, "", ""),
]

DEFAULT_SCORES = [
    ("Voyages",       5, "★★★★★"),
    ("Sorties",       4, "★★★★☆"),
    ("Loisirs",       4, "★★★★☆"),
    ("Abonnements",   3, "★★★☆☆"),
    ("Bien-être",     3, "★★★☆☆"),
    ("Shopping",      2, "★★☆☆☆"),
]


# ── Style helpers ─────────────────────────────────────────────────────────────

def _side(color=FG_DIM):
    return Side(style="thin", color=color)

def _border(color=FG_DIM):
    s = _side(color)
    return Border(left=s, right=s, top=s, bottom=s)

def _border_bottom(color=FG_DIM):
    return Border(bottom=_side(color))

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(color=FG_TEXT, size=9.5, bold=False, italic=False):
    return Font(name="Calibri", color=color, size=size, bold=bold, italic=italic)

def _align(h="left", wrap=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

def _tab_color(ws, hex_color):
    ws.sheet_properties.tabColor = hex_color

def _bg_all(ws, rows, cols, color):
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            ws.cell(r, c).fill = _fill(color)


# ── Section title band ────────────────────────────────────────────────────────

def _section_title(ws, row, col, text, fg=FG_ACCENT, bg=BG_HEADER, span=5, height=24):
    ws.row_dimensions[row].height = height
    for c in range(col, col + span):
        ws.cell(row, c).fill = _fill(bg)
    cell = ws.cell(row, col, value=text)
    cell.font = Font(name="Calibri", color=fg, bold=True, size=12)
    cell.alignment = _align("left")


# ── Column headers ────────────────────────────────────────────────────────────

def _col_headers(ws, row, headers, height=20):
    """headers = [(label, fg_color), ...]"""
    ws.row_dimensions[row].height = height
    for c, (label, fg) in enumerate(headers, 1):
        cell = ws.cell(row, c, value=label)
        cell.font = Font(name="Calibri", color=fg, bold=True, size=9)
        cell.fill = _fill("060D1A")
        cell.alignment = _align("center")
        cell.border = _border("1E3A5F")


# ── Data row ──────────────────────────────────────────────────────────────────

def _data_row(ws, row, values, colors=None, num_formats=None, height=18, alt=False):
    ws.row_dimensions[row].height = height
    bg = BG_ROW_ALT if alt else BG_ROW
    for c, val in enumerate(values, 1):
        cell = ws.cell(row, c, value=val)
        cell.fill = _fill(bg)
        fg = colors[c - 1] if colors and c - 1 < len(colors) else FG_TEXT
        cell.font = _font(fg)
        cell.alignment = _align("left")
        cell.border = _border()
        if num_formats and c - 1 < len(num_formats) and num_formats[c - 1]:
            cell.number_format = num_formats[c - 1]


def _empty_row(ws, row, ncols, height=18, alt=False):
    ws.row_dimensions[row].height = height
    bg = BG_ROW_ALT if alt else BG_ROW
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        cell.fill = _fill(bg)
        cell.font = _font(FG_TEXT)
        cell.alignment = _align()
        cell.border = _border()


def _example_row(ws, row, values, height=18):
    ws.row_dimensions[row].height = height
    for c, val in enumerate(values, 1):
        cell = ws.cell(row, c, value=val)
        cell.fill = _fill(BG_EXAMPLE)
        cell.font = Font(name="Calibri", color=FG_EXAMPLE, size=9, italic=True)
        cell.alignment = _align()
        cell.border = _border("1E3A5F")


def _total_row(ws, row, values, height=19):
    ws.row_dimensions[row].height = height
    for c, val in enumerate(values, 1):
        cell = ws.cell(row, c, value=val)
        cell.fill = _fill(BG_TOTAL)
        cell.font = Font(name="Calibri", color=FG_ACCENT, bold=True, size=9.5)
        cell.alignment = _align("right" if c > 1 else "left")
        cell.border = _border("2DD4BF")


# ── Main generator ────────────────────────────────────────────────────────────

def generate_template(month: str, savings_accounts: list[str] | None = None) -> bytes:
    year, m = map(int, month.split("-"))
    month_label = _MONTHS_FR[m - 1]

    cat_names = [name for name, _, _ in DEFAULT_CATEGORIES]
    acc_names = savings_accounts or []

    wb = Workbook()

    _build_lists(wb, cat_names, acc_names)   # hidden sheet — must be first
    _build_instructions(wb, month_label, year)
    _build_revenus(wb, month_label, year, m)
    _build_fixes(wb, month_label, year, m, cat_names)
    _build_variables(wb, month_label, year, cat_names)
    _build_categories(wb)
    _build_goals(wb, acc_names)
    _build_scores(wb)
    _build_config(wb, month, year, m)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Hidden sheet : _lists (dropdowns source) ──────────────────────────────────

def _build_lists(wb, cat_names: list, acc_names: list):
    ws = wb.create_sheet("_lists")
    ws.sheet_state = "hidden"

    ws["A1"] = "categories"
    for i, name in enumerate(cat_names, 2):
        ws.cell(i, 1, value=name)

    ws["B1"] = "comptes"
    for i, name in enumerate(acc_names, 2):
        ws.cell(i, 2, value=name)


def _list_ref(sheet, col_letter, count, header=True) -> str:
    """Returns an Excel formula reference like '_lists'!$A$2:$A$15"""
    start = 2 if header else 1
    end = start + count - 1
    return f"'_lists'!${col_letter}${start}:${col_letter}${end}"


# ── Sheet 0 : Instructions ────────────────────────────────────────────────────

def _build_instructions(wb, month_label, year):
    ws = wb.active
    ws.title = "📋 Guide"
    _tab_color(ws, FG_ACCENT)
    ws.sheet_view.showGridLines = False
    _bg_all(ws, 40, 8, "060D1A")

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 20

    def row(r, text, fg=FG_TEXT, bold=False, size=10, indent=0):
        ws.row_dimensions[r].height = 19
        cell = ws.cell(r, 2, value=("  " * indent) + text)
        cell.font = Font(name="Calibri", color=fg, bold=bold, size=size)
        cell.fill = _fill("060D1A")
        cell.alignment = _align("left")

    row(2,  f"Budget Master — Modèle d'import", FG_ACCENT, bold=True, size=15)
    row(3,  f"Période : {month_label} {year}", FG_TEXT,   bold=True, size=11)
    row(5,  "MODE D'EMPLOI",                   FG_ACCENT, bold=True, size=10)
    row(6,  "1. Remplissez chaque feuille selon vos données réelles",     FG_TEXT, indent=1)
    row(7,  "2. Ne modifiez pas les noms des feuilles ni les en-têtes",   FG_MUTED, indent=1)
    row(8,  "3. Les lignes en italique grisé sont des exemples — supprimez-les", FG_MUTED, indent=1)
    row(9,  "4. Sauvegardez au format .xlsx puis glissez dans Budget Master",    FG_MUTED, indent=1)
    row(11, "FEUILLES INCLUSES",               FG_ACCENT, bold=True, size=10)
    row(12, "💚 Revenus        — salaires, virements reçus, primes",       FG_GREEN, indent=1)
    row(13, "🔴 Dépenses fixes — loyer, abonnements, crédits récurrents",  FG_RED,   indent=1)
    row(14, "🟠 Dép. variables — courses, sorties, transports ponctuels",  FG_ORANGE,indent=1)
    row(15, "🟣 Catégories     — liste de vos catégories avec bucket 50/30/20", FG_PURPLE, indent=1)
    row(16, "🥇 Objectifs      — vos projets d'épargne à atteindre",       FG_GOLD,  indent=1)
    row(17, "⭐ Scores wants   — priorité de vos dépenses plaisir",        FG_ACCENT,indent=1)
    row(19, "COLONNES CLÉS",                   FG_ACCENT, bold=True, size=10)
    row(20, "Libellé      — nom de la transaction (ex: Loyer, Netflix)",   FG_TEXT, indent=1)
    row(21, "Montant      — valeur positive en euros",                      FG_TEXT, indent=1)
    row(22, "Catégorie    — doit correspondre à un nom de la feuille Catégories", FG_TEXT, indent=1)
    row(23, "Jour         — jour du mois (1-31) de la transaction",        FG_TEXT, indent=1)
    row(24, "Occurrences  — nb de fois dans le mois (dép. variables)",     FG_TEXT, indent=1)
    row(25, "Bucket       — needs / wants / savings (feuille Catégories)", FG_TEXT, indent=1)
    row(26, "Score        — pertinence de 1 (faible) à 5 (essentiel)",     FG_TEXT, indent=1)
    row(28, f"⚠️  Ce fichier est configuré pour : {month_label} {year}", FG_ORANGE, bold=True)


# ── Sheet 1 : Revenus ─────────────────────────────────────────────────────────

def _build_revenus(wb, month_label, year, m):
    ws = wb.create_sheet("💚 Revenus")
    _tab_color(ws, FG_GREEN)
    ws.sheet_view.showGridLines = False
    _bg_all(ws, 55, 5, BG_PAGE)

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14

    _section_title(ws, 1, 1, f"💚 Revenus — {month_label} {year}", FG_GREEN, span=3)
    _col_headers(ws, 2, [
        ("Source / Libellé", FG_GREEN),
        ("Montant (€)",      FG_GREEN),
        ("Jour du mois",     FG_MUTED),
    ])

    # Pre-filled examples (user's real data)
    examples = [
        ("Virement DCS'Easyware", 1699, 1),
        ("Virement Perce-Neige",  1300, 1),
    ]
    for i, (label, amt, day) in enumerate(examples):
        _data_row(ws, 3 + i, [label, amt, day],
                  colors=[FG_TEXT, FG_GREEN, FG_MUTED],
                  num_formats=[None, '#,##0.00 €', None],
                  alt=(i % 2 == 1))

    for r in range(3 + len(examples), 35):
        _empty_row(ws, r, 3, alt=(r % 2 == 0))

    # Total row
    last_data = 3 + len(examples) - 1
    _total_row(ws, 35, ["TOTAL", f"=SUM(B3:B{last_data})", "—"])

    # Validations
    dv_day = DataValidation(type="whole", operator="between", formula1="1", formula2="31",
                            showErrorMessage=True, errorTitle="Jour invalide",
                            error="Entrez un jour entre 1 et 31.")
    ws.add_data_validation(dv_day); dv_day.add("C3:C34")

    dv_amt = DataValidation(type="decimal", operator="greaterThan", formula1="0",
                            showErrorMessage=True, errorTitle="Montant invalide",
                            error="Le montant doit être positif.")
    ws.add_data_validation(dv_amt); dv_amt.add("B3:B34")

    ws.freeze_panes = "A3"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True


# ── Sheet 2 : Dépenses fixes ──────────────────────────────────────────────────

def _build_fixes(wb, month_label, year, m, cat_names=None):
    ws = wb.create_sheet("🔴 Dépenses fixes")
    _tab_color(ws, FG_RED)
    ws.sheet_view.showGridLines = False
    _bg_all(ws, 55, 6, BG_PAGE)

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 14

    _section_title(ws, 1, 1, f"🔴 Dépenses fixes — {month_label} {year}", FG_RED, span=4)
    _col_headers(ws, 2, [
        ("Libellé",      FG_RED),
        ("Montant (€)",  FG_RED),
        ("Catégorie",    FG_MUTED),
        ("Jour du mois", FG_MUTED),
    ])

    # Pre-filled with user's real data
    examples = [
        ("Loyer",       719.00,  "Logement", 1),
        ("Électricité", 67.00,   "Énergie",  3),
        ("Cetelem",     179.92,  "Crédit",   5),
        ("Floa Bank",   113.00,  "Crédit",   5),
        ("Cofidis",     190.26,  "Crédit",   5),
        ("BNP",         180.00,  "Crédit",   5),
    ]
    for i, (label, amt, cat, day) in enumerate(examples):
        _data_row(ws, 3 + i, [label, amt, cat, day],
                  colors=[FG_TEXT, FG_RED, FG_MUTED, FG_MUTED],
                  num_formats=[None, '#,##0.00 €', None, None],
                  alt=(i % 2 == 1))

    for r in range(3 + len(examples), 45):
        _empty_row(ws, r, 4, alt=(r % 2 == 0))

    last_data = 3 + len(examples) - 1
    _total_row(ws, 45, ["TOTAL", f"=SUM(B3:B{last_data})", "—", "—"])

    dv_day = DataValidation(type="whole", operator="between", formula1="1", formula2="31",
                            showErrorMessage=True)
    ws.add_data_validation(dv_day); dv_day.add("D3:D44")

    if cat_names:
        dv_cat = DataValidation(type="list", formula1=_list_ref("_lists", "A", len(cat_names)),
                                showDropDown=False, showErrorMessage=True,
                                errorTitle="Catégorie invalide", error="Choisissez une catégorie de la liste.")
        ws.add_data_validation(dv_cat); dv_cat.add("C3:C44")

    ws.freeze_panes = "A3"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True


# ── Sheet 3 : Dépenses variables ──────────────────────────────────────────────

def _build_variables(wb, month_label, year, cat_names=None):
    ws = wb.create_sheet("🟠 Dép. variables")
    _tab_color(ws, FG_ORANGE)
    ws.sheet_view.showGridLines = False
    _bg_all(ws, 55, 7, BG_PAGE)

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 18

    _section_title(ws, 1, 1, f"🟠 Dépenses variables — {month_label} {year}", FG_ORANGE, span=5)
    _col_headers(ws, 2, [
        ("Libellé",            FG_ORANGE),
        ("Min (€)",            FG_ORANGE),
        ("Max (€)",            FG_ORANGE),
        ("Catégorie",          FG_MUTED),
        ("Occurrences / mois", FG_MUTED),
    ])

    # Pre-filled with user's real data
    examples = [
        ("Collation mois",   150, 200, "Alimentation", 3),
        ("Navigo",            90,  94, "Transport",    2),
        ("Pharmacie",         15,  55, "Santé",        1),
        ("Claude AI",         17,  24, "Abonnements",  1),
        ("Restaurant",        30,  70, "Sorties",      2),
        ("Café / Snack",       8,  25, "Sorties",      3),
        ("Vêtements",         70, 100, "Shopping",     1),
        ("Loisirs / Sorties", 100,150, "Loisirs",      1),
        ("Coiffeur",          30,  45, "Bien-être",    1),
        ("Épargne Congo",    200, 300, "Voyages",      1),
    ]
    for i, row_vals in enumerate(examples):
        _data_row(ws, 3 + i, list(row_vals),
                  colors=[FG_TEXT, FG_ORANGE, FG_ORANGE, FG_MUTED, FG_MUTED],
                  num_formats=[None, '#,##0.00 €', '#,##0.00 €', None, None],
                  alt=(i % 2 == 1))

    for r in range(3 + len(examples), 45):
        _empty_row(ws, r, 5, alt=(r % 2 == 0))

    dv_occ = DataValidation(type="whole", operator="between", formula1="1", formula2="10",
                            showErrorMessage=True, errorTitle="Occurrences",
                            error="Entrez un nombre entre 1 et 10.")
    ws.add_data_validation(dv_occ); dv_occ.add("E3:E44")

    if cat_names:
        dv_cat = DataValidation(type="list", formula1=_list_ref("_lists", "A", len(cat_names)),
                                showDropDown=False, showErrorMessage=True,
                                errorTitle="Catégorie invalide", error="Choisissez une catégorie de la liste.")
        ws.add_data_validation(dv_cat); dv_cat.add("D3:D44")

    ws.freeze_panes = "A3"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True


# ── Sheet 4 : Catégories ──────────────────────────────────────────────────────

def _build_categories(wb):
    ws = wb.create_sheet("🟣 Catégories")
    _tab_color(ws, FG_PURPLE)
    ws.sheet_view.showGridLines = False
    _bg_all(ws, 55, 5, BG_PAGE)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 30

    _section_title(ws, 1, 1, "🟣 Catégories — Buckets 50/30/20", FG_PURPLE, span=4)

    # Sub-headers explaining buckets
    ws.row_dimensions[2].height = 16
    for c, (txt, fg) in enumerate([
        ("", FG_MUTED), ("needs = besoins", FG_BLUE),
        ("", FG_MUTED), ("wants = plaisirs | savings = épargne", FG_MUTED)
    ], 1):
        cell = ws.cell(2, c, value=txt)
        cell.font = Font(name="Calibri", color=fg, size=8, italic=True)
        cell.fill = _fill("08101E")
        cell.alignment = _align()

    _col_headers(ws, 3, [
        ("Nom",             FG_PURPLE),
        ("Bucket",          FG_PURPLE),
        ("Icône",           FG_MUTED),
        ("Description (optionnel)", FG_DIM),
    ])

    bucket_colors = {"needs": FG_BLUE, "wants": FG_ORANGE, "savings": FG_GREEN}

    for i, (name, bucket, icon) in enumerate(DEFAULT_CATEGORIES):
        fg_bucket = bucket_colors.get(bucket, FG_MUTED)
        _data_row(ws, 4 + i, [name, bucket, icon, ""],
                  colors=[FG_TEXT, fg_bucket, FG_TEXT, FG_MUTED],
                  alt=(i % 2 == 1))

    for r in range(4 + len(DEFAULT_CATEGORIES), 45):
        _empty_row(ws, r, 4, alt=(r % 2 == 0))

    # Bucket validation dropdown
    dv_bucket = DataValidation(type="list", formula1='"needs,wants,savings"',
                               showErrorMessage=True, errorTitle="Bucket invalide",
                               error="Choisissez : needs, wants ou savings")
    ws.add_data_validation(dv_bucket)
    dv_bucket.add(f"B4:B44")

    ws.freeze_panes = "A4"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True


# ── Sheet 5 : Objectifs ───────────────────────────────────────────────────────

def _build_goals(wb, acc_names=None):
    ws = wb.create_sheet("🥇 Objectifs")
    _tab_color(ws, FG_GOLD)
    ws.sheet_view.showGridLines = False
    _bg_all(ws, 35, 8, BG_PAGE)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 22

    _section_title(ws, 1, 1, "🥇 Objectifs d'épargne", FG_GOLD, span=7)
    _col_headers(ws, 2, [
        ("Objectif",           FG_GOLD),
        ("Icône",              FG_MUTED),
        ("Actuel (€)",         FG_GREEN),
        ("Cible (€)",          FG_GOLD),
        ("Progression",        FG_ACCENT),
        ("Deadline",           FG_MUTED),
        ("Compte épargne lié", FG_BLUE),
    ])

    for i, (name, icon, current, target, deadline, savings_account) in enumerate(DEFAULT_GOALS):
        pct = f"=IF(D{3+i}>0,C{3+i}/D{3+i},0)"
        _data_row(ws, 3 + i, [name, icon, current, target, pct, deadline, savings_account],
                  colors=[FG_TEXT, FG_TEXT, FG_GREEN, FG_GOLD, FG_ACCENT, FG_MUTED, FG_BLUE],
                  num_formats=[None, None, '#,##0.00 €', '#,##0.00 €', '0%', 'DD/MM/YYYY', None],
                  alt=(i % 2 == 1))

    for r in range(3 + len(DEFAULT_GOALS), 25):
        _empty_row(ws, r, 7, alt=(r % 2 == 0))

    # Amount validations
    dv_amt = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0",
                            showErrorMessage=True)
    ws.add_data_validation(dv_amt); dv_amt.add("C3:D24")

    if acc_names:
        dv_acc = DataValidation(type="list", formula1=_list_ref("_lists", "B", len(acc_names)),
                                showDropDown=False, showErrorMessage=False)
        ws.add_data_validation(dv_acc); dv_acc.add("G3:G24")

    ws.freeze_panes = "A3"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True


# ── Sheet 6 : Scores wants ────────────────────────────────────────────────────

def _build_scores(wb):
    ws = wb.create_sheet("⭐ Scores wants")
    _tab_color(ws, FG_ACCENT)
    ws.sheet_view.showGridLines = False
    _bg_all(ws, 35, 5, BG_PAGE)

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 34

    _section_title(ws, 1, 1, "⭐ Scores de pertinence — Bucket Wants", FG_ACCENT, span=4)

    ws.row_dimensions[2].height = 16
    hint = ws.cell(2, 1, value="Score de 1 (dépense peu prioritaire) à 5 (essentielle) — influence le budget alloué")
    hint.font = Font(name="Calibri", color=FG_MUTED, size=8, italic=True)
    hint.fill = _fill("08101E")
    hint.alignment = _align()

    _col_headers(ws, 3, [
        ("Catégorie",   FG_ACCENT),
        ("Score (/5)",  FG_ACCENT),
        ("Étoiles",     FG_GOLD),
        ("Conseil",     FG_DIM),
    ])

    advice = {
        5: "Dépense quasi-incontournable pour votre qualité de vie",
        4: "Importante, limitez sans éliminer",
        3: "Neutre — à surveiller",
        2: "Confort non essentiel — premier poste à réduire",
        1: "Optionnel — éliminer en cas de besoin",
    }
    star_map = {5: "★★★★★", 4: "★★★★☆", 3: "★★★☆☆", 2: "★★☆☆☆", 1: "★☆☆☆☆"}

    score_colors = {5: FG_GREEN, 4: FG_ACCENT, 3: FG_BLUE, 2: FG_ORANGE, 1: FG_RED}

    for i, (cat, score, _) in enumerate(DEFAULT_SCORES):
        stars = star_map.get(score, "★★★☆☆")
        tip   = advice.get(score, "")
        fg_s  = score_colors.get(score, FG_MUTED)
        _data_row(ws, 4 + i, [cat, score, stars, tip],
                  colors=[FG_TEXT, fg_s, FG_GOLD, FG_MUTED],
                  alt=(i % 2 == 1))

    for r in range(4 + len(DEFAULT_SCORES), 25):
        _empty_row(ws, r, 4, alt=(r % 2 == 0))

    dv_score = DataValidation(type="whole", operator="between", formula1="1", formula2="5",
                              showErrorMessage=True, errorTitle="Score invalide",
                              error="Entrez un score de 1 à 5.")
    ws.add_data_validation(dv_score); dv_score.add("B4:B24")

    ws.freeze_panes = "A4"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True


# ── Hidden config ─────────────────────────────────────────────────────────────

def _build_config(wb, month, year, m):
    ws = wb.create_sheet("_config")
    ws.sheet_state = "hidden"
    ws["A1"] = "BudgetMaster-Template-v2"
    ws["A2"] = month
    ws["A3"] = year
    ws["A4"] = m
